"""XLSX parser — extracts structured sections and tables from Excel spreadsheets.

Requires the ``xlsx`` extra::

    pip install openpyxl
"""

from __future__ import annotations

from pathlib import Path

from smartchunk.models import DocumentSection, TableData
from smartchunk.parsers.base import BaseParser


class XlsxParser(BaseParser):
    """Parses Microsoft Excel (.xlsx) documents."""

    supported_extensions = [".xlsx"]

    def parse(self, filepath: str | Path) -> list[DocumentSection]:
        filepath = Path(filepath)

        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError as exc:
            raise ImportError(
                "XLSX parsing requires openpyxl. Install with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        sections: list[DocumentSection] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            rows_data: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                # Clean elements to string
                cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
                # Only append if not completely empty row
                if any(cleaned_row):
                    rows_data.append(cleaned_row)

            if not rows_data:
                continue

            # First row is headers, rest is data rows
            headers = rows_data[0]
            data_rows = rows_data[1:]

            table_data = TableData(headers=headers, rows=data_rows, caption=sheet_name)

            sections.append(
                DocumentSection(
                    text=table_data.to_markdown(),
                    heading=sheet_name,
                    level=1,
                    parent_headings=[],
                    content_type="table",
                    table=table_data,
                    metadata={"source": filepath.name, "sheet_name": sheet_name},
                )
            )

        return sections
